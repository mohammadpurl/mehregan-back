"""خواندن فایل اکسل اقلام هزینه تنخواه."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.schemas.petty_cash import PettyCashExpenseLineIn

HEADER_ALIASES: dict[str, set[str]] = {
    "description": {
        "description",
        "desc",
        "شرح",
        "توضیح",
        "عنوان",
        "شرح هزینه",
    },
    "amount": {
        "amount",
        "مبلغ",
        "مبلغ (ریال)",
        "مبلغ(ریال)",
        "price",
    },
    "expense_date": {
        "expense_date",
        "expensedate",
        "date",
        "تاریخ",
        "تاریخ هزینه",
    },
}


def _norm_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "").replace("_", "")


def _map_headers(row_cells) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row_cells):
        raw = _norm_header(cell.value if cell else None)
        if not raw:
            continue
        for field, aliases in HEADER_ALIASES.items():
            norm_aliases = {a.replace(" ", "").replace("_", "") for a in aliases}
            if raw in norm_aliases and field not in mapping:
                mapping[field] = idx
    return mapping


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_expense_excel(content: bytes) -> list[PettyCashExpenseLineIn]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("فایل اکسل خالی است")

    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        raise ValueError("فایل اکسل خالی است")

    header_map = _map_headers(rows[0])
    if "description" not in header_map or "amount" not in header_map:
        raise ValueError(
            "ستون‌های الزامی یافت نشد: حداقل «شرح» و «مبلغ» در سطر اول"
        )

    lines: list[PettyCashExpenseLineIn] = []
    for row in rows[1:]:
        if not row:
            continue
        desc_cell = row[header_map["description"]] if header_map["description"] < len(row) else None
        amt_cell = row[header_map["amount"]] if header_map["amount"] < len(row) else None
        desc = (desc_cell.value if desc_cell else None) or ""
        desc = str(desc).strip()
        if not desc:
            continue
        raw_amount = amt_cell.value if amt_cell else None
        if raw_amount is None or str(raw_amount).strip() == "":
            continue
        try:
            amount = float(raw_amount)
        except (TypeError, ValueError):
            raise ValueError(f"مبلغ نامعتبر در ردیف: {desc}") from None
        if amount <= 0:
            continue

        expense_date = None
        if "expense_date" in header_map and header_map["expense_date"] < len(row):
            expense_date = _parse_date(row[header_map["expense_date"]].value)

        lines.append(
            PettyCashExpenseLineIn(
                description=desc,
                amount=amount,
                expense_date=expense_date,
            )
        )

    wb.close()
    if not lines:
        raise ValueError("هیچ ردیف هزینه معتبری در فایل اکسل یافت نشد")
    return lines
